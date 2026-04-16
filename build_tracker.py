from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "MVP Tracker"

# ── helpers ──────────────────────────────────────────────────────────────────
def solid(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def font(bold=False, italic=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, italic=italic, size=size, color=color)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── status → row fill color ───────────────────────────────────────────────────
STATUS_FILL = {
    "✅ Done":           "d1fae5",
    "🔧 Needs Config":  "fef9c3",
    "⚠️ Partial":       "ffedd5",
    "❌ Not Started":   "fee2e2",
}

# ── data ──────────────────────────────────────────────────────────────────────
rows = [
    # (status, area, task, notes, priority)
    ("✅ Done",          "Landing Page",    "Navbar (logo, nav links, phone, CTA)",              "Fully built and responsive",                                                           "Medium"),
    ("✅ Done",          "Landing Page",    "Hero section (headline, image, CTAs)",               "Stock image from Unsplash",                                                            "Medium"),
    ("✅ Done",          "Landing Page",    "Stats bar (24/7, 100% screened, <24h)",              "Static values",                                                                        "Low"),
    ("✅ Done",          "Landing Page",    "About section (text + image)",                       "Copy and image in place",                                                              "Low"),
    ("✅ Done",          "Landing Page",    "Services section (4 service cards)",                 "Temporary, Direct Hire, Per Diem, Locum Tenens",                                       "Low"),
    ("✅ Done",          "Landing Page",    "Why Choose Us section",                              "3 feature cards + mission/vision",                                                     "Low"),
    ("✅ Done",          "Landing Page",    "Contact bar (phone, email, address)",                "Real contact details in code",                                                         "Low"),
    ("✅ Done",          "Landing Page",    "Footer with links",                                  "Copyright year says 2024 — needs updating",                                            "Low"),
    ("✅ Done",          "Form — Apply",    "Form UI (name, email, phone, message)",              "Styled, responsive",                                                                   "Medium"),
    ("✅ Done",          "Form — Apply",    "CV file upload (PDF/DOC/DOCX, max 10MB)",            "Drag-style button, file preview",                                                      "Medium"),
    ("✅ Done",          "Form — Apply",    "Client-side validation + inline errors",             "All fields validated on submit & blur",                                                "Medium"),
    ("✅ Done",          "Form — Apply",    "Loading / success / error states",                   "Spinner, success card, error banner",                                                  "Medium"),
    ("✅ Done",          "Form — Find Staff","Form UI (org, contact, role, headcount, date)",     "All fields built",                                                                     "Medium"),
    ("✅ Done",          "Form — Find Staff","Client-side validation + inline errors",            "Required fields validated",                                                            "Medium"),
    ("✅ Done",          "Form — Find Staff","Loading / success / error states",                  "Same pattern as apply form",                                                           "Medium"),
    ("✅ Done",          "Backend",         "Express server with /api/apply endpoint",            "Accepts multipart form + CV attachment",                                               "High"),
    ("✅ Done",          "Backend",         "Express server with /api/request-staff endpoint",    "Sends JSON email to recipient",                                                        "High"),
    ("✅ Done",          "Backend",         "Multer file handling (CV attachment)",               "Stores in memory, attaches to email",                                                  "High"),
    ("✅ Done",          "Backend",         "Nodemailer email sending logic",                     "Both endpoints wired up",                                                              "High"),
    ("🔧 Needs Config",  "Backend",         "SMTP credentials in .env file",                      ".env exists but has placeholder values — must fill in real host, user, password",      "Critical"),
    ("🔧 Needs Config",  "Backend",         "RECIPIENT_EMAIL in .env",                            'Set to "hiring@yourdomain.com" — must be customer\'s real inbox',                     "Critical"),
    ("⚠️ Partial",       "Backend",         "Start the Express server before the site works",     'Run "npm run server" in a second terminal — or add "concurrently" to run both at once',"Critical"),
    ("⚠️ Partial",       "Backend",         "CORS origin hardcoded to localhost:5173",            "Will break in production — must update to real domain before deploying",               "High"),
    ("⚠️ Partial",       "Frontend",        "API URLs hardcoded to http://localhost:3001",         "Two fetch() calls in App.jsx point to localhost — must change to real server URL for production", "High"),
    ("❌ Not Started",   "Deployment",      "Choose and set up hosting for the frontend",         "Recommended: Vercel or Netlify (free)",                                                "Critical"),
    ("❌ Not Started",   "Deployment",      "Choose and set up hosting for the Express backend",  "Recommended: Render free tier or Railway (~$5/mo)",                                    "Critical"),
    ("❌ Not Started",   "Deployment",      "Buy and connect a domain name",                      "Recommended: Namecheap (~$10–15/year)",                                                "High"),
    ("❌ Not Started",   "Deployment",      "Update CORS + API URLs for production domain",       "After hosting is chosen, update both files",                                           "High"),
    ("❌ Not Started",   "Email",           "Test end-to-end email delivery",                     "Submit both forms, confirm emails arrive with CV attached",                            "Critical"),
    ("❌ Not Started",   "Polish",          "Update footer copyright year from 2024 to 2025",     "Minor but looks unprofessional",                                                       "Low"),
]

# ── Row 1: Title ───────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
ws["A1"] = "Tender Smiles Healthcare Staffing — MVP Tracker"
ws["A1"].font      = font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill      = solid("1e293b")
ws["A1"].alignment = center
ws.row_dimensions[1].height = 36

# ── Row 2: Subtitle ────────────────────────────────────────────────────────────
ws.merge_cells("A2:F2")
ws["A2"] = "Last updated: 16 April 2026"
ws["A2"].font      = font(italic=True, size=10, color="64748b")
ws["A2"].fill      = solid("FFFFFF")
ws["A2"].alignment = center
ws.row_dimensions[2].height = 20

# ── Row 3: Spacer ──────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 8

# ── Row 4: Column headers ──────────────────────────────────────────────────────
headers = ["#", "Area", "Task / Item", "Status", "Notes", "Priority"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font      = font(bold=True, color="FFFFFF")
    cell.fill      = solid("532cd8")
    cell.alignment = center
ws.row_dimensions[4].height = 22

# ── Data rows (starting at row 5) ─────────────────────────────────────────────
DATA_START = 5
for i, (status, area, task, notes, priority) in enumerate(rows, start=1):
    r = DATA_START + i - 1
    row_fill = solid(STATUS_FILL[status])
    values = [i, area, task, status, notes, priority]
    aligns = [center, left, left, center, left, center]
    for col, (val, aln) in enumerate(zip(values, aligns), start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font      = font()
        cell.fill      = row_fill
        cell.alignment = aln
    ws.row_dimensions[r].height = 18

DATA_END = DATA_START + len(rows) - 1  # row 34

# ── Summary table (2 blank rows after data) ────────────────────────────────────
SUM_START = DATA_END + 3  # row 37

# Title row
ws.merge_cells(f"A{SUM_START}:B{SUM_START}")
ws[f"A{SUM_START}"] = "Summary"
ws[f"A{SUM_START}"].font      = font(bold=True, color="FFFFFF")
ws[f"A{SUM_START}"].fill      = solid("532cd8")
ws[f"A{SUM_START}"].alignment = center
ws.row_dimensions[SUM_START].height = 18

# Header row
hdr_fill = solid("e2e8f0")
for col, val in enumerate(["Status", "Count"], start=1):
    cell = ws.cell(row=SUM_START + 1, column=col, value=val)
    cell.font      = font(bold=True)
    cell.fill      = hdr_fill
    cell.alignment = center
ws.row_dimensions[SUM_START + 1].height = 16

# COUNTIF rows — column D holds status values (rows 5:34)
count_rows = [
    ("✅ Done",          f'=COUNTIF(D{DATA_START}:D{DATA_END},"✅ Done")'),
    ("🔧 Needs Config",  f'=COUNTIF(D{DATA_START}:D{DATA_END},"🔧 Needs Config")'),
    ("⚠️ Partial",       f'=COUNTIF(D{DATA_START}:D{DATA_END},"⚠️ Partial")'),
    ("❌ Not Started",   f'=COUNTIF(D{DATA_START}:D{DATA_END},"❌ Not Started")'),
]

count_cells = []
for offset, (label, formula) in enumerate(count_rows, start=2):
    r = SUM_START + offset
    ws.cell(row=r, column=1, value=label).font      = font()
    ws.cell(row=r, column=1).fill                   = solid("FFFFFF")
    ws.cell(row=r, column=1).alignment              = left
    count_cell = ws.cell(row=r, column=2, value=formula)
    count_cell.font      = font()
    count_cell.fill      = solid("FFFFFF")
    count_cell.alignment = center
    count_cells.append(f"B{r}")
    ws.row_dimensions[r].height = 16

# Total row
total_row = SUM_START + 2 + len(count_rows)
ws.cell(row=total_row, column=1, value="Total").font      = font(bold=True)
ws.cell(row=total_row, column=1).fill                     = solid("e2e8f0")
ws.cell(row=total_row, column=1).alignment                = left
sum_formula = f"=SUM({count_cells[0]}:{count_cells[-1]})"
tot = ws.cell(row=total_row, column=2, value=sum_formula)
tot.font      = font(bold=True)
tot.fill      = solid("e2e8f0")
tot.alignment = center
ws.row_dimensions[total_row].height = 16

# ── Column widths ──────────────────────────────────────────────────────────────
col_widths = [5, 22, 42, 14, 40, 12]
for col, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Save ───────────────────────────────────────────────────────────────────────
out = "/Users/max/projects/healthcare-staffing/MVP_Tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Data rows: {DATA_START}–{DATA_END}, Summary starts: {SUM_START}")
